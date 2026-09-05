"""学校の年間行事予定を、内部形式（events 配列）に正規化して取り込む。

学校ごとに行事予定表の形式が違うため、ここでは「列の意味を推測して読む」ことに徹する。
対応形式:
  - CSV / TSV（日付列 + 行事名列）
  - Excel（.xlsx / .xlsm）: 縦持ち（日付・行事の列）と、横持ち（列=月、行=日）の両方

読み取れなかった行は捨てずに warnings として返す。人が確認・修正する前提の設計。
完全自動を狙わず、「8割を機械が読んで、残りを人が直す」ための部品。
"""

from __future__ import annotations

import csv
import datetime as _dt
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable

__all__ = ["ImportResult", "import_events", "parse_loose_date"]

_DATE_HEADERS = ("日付", "月日", "date", "日にち", "年月日")
_TITLE_HEADERS = ("行事", "予定", "内容", "行事名", "title", "event", "備考欄")
_NOTE_HEADERS = ("備考", "note", "メモ", "下校", "日課")

# 「4/8」「4月8日」「2025-04-08」「4.8」などを拾う
_MD_RE = re.compile(r"(?P<m>\d{1,2})\s*[/月.\-]\s*(?P<d>\d{1,2})")
_YMD_RE = re.compile(r"(?P<y>\d{4})\s*[-/年]\s*(?P<m>\d{1,2})\s*[-/月]\s*(?P<d>\d{1,2})")


@dataclass
class ImportResult:
    """取り込み結果。events はそのまま入力JSONに入れられる形。"""

    events: list[dict[str, str]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def count(self) -> int:
        return len(self.events)


def parse_loose_date(value: Any, school_year: int) -> _dt.date | None:
    """ゆるい書式の日付を解釈する。年が無い場合は年度から推測する。

    月が1〜3月なら翌年、4〜12月ならその年とみなす（学校年度の慣習）。
    """
    if value is None or value == "":
        return None
    if isinstance(value, _dt.datetime):
        return value.date()
    if isinstance(value, _dt.date):
        return value

    text = str(value).strip()
    if not text:
        return None

    ymd = _YMD_RE.search(text)
    if ymd:
        try:
            return _dt.date(int(ymd["y"]), int(ymd["m"]), int(ymd["d"]))
        except ValueError:
            return None

    md = _MD_RE.search(text)
    if md:
        month, day = int(md["m"]), int(md["d"])
        if not (1 <= month <= 12):
            return None
        year = school_year + 1 if month <= 3 else school_year
        try:
            return _dt.date(year, month, day)
        except ValueError:
            return None
    return None


def _match_header(headers: Iterable[str], candidates: tuple[str, ...]) -> int | None:
    for i, header in enumerate(headers):
        text = str(header or "").strip().lower()
        if any(c.lower() in text for c in candidates):
            return i
    return None


def _rows_from_csv(path: Path) -> list[list[Any]]:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    delimiter = "\t" if path.suffix.lower() == ".tsv" or text.count("\t") > text.count(",") else ","
    return [row for row in csv.reader(text.splitlines(), delimiter=delimiter)]


def _rows_from_excel(path: Path) -> list[list[Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _import_vertical(rows: list[list[Any]], school_year: int, result: ImportResult) -> bool:
    """縦持ち（1行1行事、日付列と行事名列がある）形式として読む。"""
    if not rows:
        return False

    header_index = None
    for i, row in enumerate(rows[:10]):
        cells = [str(c or "") for c in row]
        if _match_header(cells, _DATE_HEADERS) is not None and _match_header(cells, _TITLE_HEADERS) is not None:
            header_index = i
            break
    if header_index is None:
        return False

    header = [str(c or "") for c in rows[header_index]]
    date_col = _match_header(header, _DATE_HEADERS)
    title_col = _match_header(header, _TITLE_HEADERS)
    note_col = _match_header(header, _NOTE_HEADERS)
    assert date_col is not None and title_col is not None

    for line_no, row in enumerate(rows[header_index + 1 :], start=header_index + 2):
        if not row or all(c in (None, "") for c in row):
            continue
        raw_date = row[date_col] if date_col < len(row) else None
        title = str(row[title_col] or "").strip() if title_col < len(row) else ""
        if not title:
            continue
        date = parse_loose_date(raw_date, school_year)
        if date is None:
            result.warnings.append(f"{line_no}行目: 日付 '{raw_date}' を解釈できませんでした（行事: {title}）")
            continue
        event = {"date": date.isoformat(), "title": title}
        if note_col is not None and note_col < len(row) and row[note_col]:
            event["note"] = str(row[note_col]).strip()
        result.events.append(event)
    return True


def _import_matrix(rows: list[list[Any]], school_year: int, result: ImportResult) -> bool:
    """横持ち（列=月、行=日）形式として読む。現行 Excel の「行事予定(一覧)」がこの形。"""
    if not rows:
        return False

    # 月ヘッダの行を探す（"4月" や 4 が横に並んでいる行）
    month_row_index = None
    month_cols: dict[int, int] = {}
    for i, row in enumerate(rows[:12]):
        found: dict[int, int] = {}
        for col, cell in enumerate(row):
            text = str(cell or "").strip()
            m = re.fullmatch(r"(\d{1,2})\s*月?", text)
            if m and 1 <= int(m.group(1)) <= 12:
                found[col] = int(m.group(1))
        if len(found) >= 6:  # 半年分以上並んでいれば月ヘッダとみなす
            month_row_index = i
            month_cols = found
            break
    if month_row_index is None:
        return False

    for row in rows[month_row_index + 1 :]:
        if not row:
            continue
        for col, month in month_cols.items():
            if col >= len(row):
                continue
            cell = row[col]
            # セル自体が日付なら隣のセルを行事名とみなす
            date = parse_loose_date(cell, school_year)
            title = ""
            if date is None:
                continue
            if col + 1 < len(row) and row[col + 1]:
                title = str(row[col + 1]).strip()
            if not title:
                continue
            result.events.append({"date": date.isoformat(), "title": title})
    return bool(result.events)


def import_events(path: str | Path, school_year: int) -> ImportResult:
    """行事予定ファイルを読み、events 配列に正規化する。

    形式を自動判定し、読めなかった行は warnings に残す。
    """
    p = Path(path)
    result = ImportResult()
    if not p.exists():
        result.warnings.append(f"ファイルが見つかりません: {p}")
        return result

    suffix = p.suffix.lower()
    try:
        if suffix in (".csv", ".tsv", ".txt"):
            rows = _rows_from_csv(p)
        elif suffix in (".xlsx", ".xlsm", ".xltx"):
            rows = _rows_from_excel(p)
        else:
            result.warnings.append(
                f"未対応の形式です: {suffix}。CSV か Excel(.xlsx/.xlsm) に変換してください"
            )
            return result
    except Exception as exc:  # ファイル破損など
        result.warnings.append(f"ファイルを読めませんでした: {exc}")
        return result

    if _import_vertical(rows, school_year, result):
        pass
    elif _import_matrix(rows, school_year, result):
        pass
    else:
        result.warnings.append(
            "行事予定の形式を判別できませんでした。"
            "1列目に日付、2列目に行事名を置いた CSV（1行目に『日付,行事』の見出し）にしてください"
        )

    # 同じ日付・同じ行事名の重複を除去
    seen: set[tuple[str, str]] = set()
    unique: list[dict[str, str]] = []
    for event in result.events:
        key = (event["date"], event["title"])
        if key in seen:
            continue
        seen.add(key)
        unique.append(event)
    result.events = sorted(unique, key=lambda e: (e["date"], e["title"]))
    return result
